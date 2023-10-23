from time import  strftime
from openpyxl import load_workbook

def getdirectorypath():
    from Tkinter import Tk
    from tkFileDialog import askdirectory

    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    filename = askdirectory()
    return filename

def getfilename():
    from Tkinter import Tk
    from tkFileDialog import askopenfilename

    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    myFormats = [
    ('Excel Workbook','*.xlsx '),
    ('Excel 97- Excel 2003 Workbook','*.xls '),
    ('Excel Workbook (code)','*.xlsm '),
    ('Excel 97- Excel 2003 Workbook(code)','*.xlm'),
    ]
    filename = askopenfilename(title='Select your Xls file for Static Register Evaluation ', filetypes=myFormats)
    return filename


DataToAppendStart='<?xml version="1.0" encoding="UTF-8" standalone="no"?>\
<AUTOSAR xmlns="http://autosar.org/schema/r4.0" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="http://autosar.org/schema/r4.0 AUTOSAR_4-0-3.xsd">\
  <AR-PACKAGES>\n\
    <AR-PACKAGE UUID="B7EDD042-4044-414C-9970-1C2313A88AC7">\n\
      <SHORT-NAME>ActiveEcuC</SHORT-NAME>\n\
      <ELEMENTS>\n\
        <ECUC-MODULE-CONFIGURATION-VALUES UUID="c4fafef7-d0ee-4639-9c36-3969fdda2374">\n\
          <SHORT-NAME>VrfyCritReg</SHORT-NAME>\n\
          <DEFINITION-REF DEST="ECUC-MODULE-DEF">/Nxtr/VrfyCritReg</DEFINITION-REF>\n\
          <IMPLEMENTATION-CONFIG-VARIANT>VARIANT-PRE-COMPILE</IMPLEMENTATION-CONFIG-VARIANT>\n\
          <CONTAINERS>\n'

DataToEdit8BitInit ='<ECUC-CONTAINER-VALUE UUID="224569eb-a267-46de-b306-e4db22625bf1">\n\
              <SHORT-NAME>SIGNALNAME8BIT</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitInitSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitInitSignallist/Address</DEFINITION-REF>\n\
                  <VALUE>ADDRESS8BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitInitSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>VALUE8BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitInitSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>MASK8BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'
DataToEdit16BitInit ='<ECUC-CONTAINER-VALUE UUID="224569eb-a267-46de-b306-e4db22625bf1">\n\
              <SHORT-NAME>SIGNALNAME16BIT</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitInitSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitInitSignallist/Address</DEFINITION-REF>\n\
                  <VALUE>ADDRESS16BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitInitSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>VALUE16BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitInitSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>MASK16BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'
DataToEdit32BitInit ='<ECUC-CONTAINER-VALUE UUID="224569eb-a267-46de-b306-e4db22625bf1">\n\
              <SHORT-NAME>SIGNALNAME32BIT</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitInitSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitInitSignallist/Address</DEFINITION-REF>\n\
                  <VALUE>ADDRESS32BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitInitSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>VALUE32BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitInitSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>MASK32BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'

DataToEdit8BitPer ='<ECUC-CONTAINER-VALUE UUID="224569eb-a267-46de-b306-e4db22625bf1">\n\
              <SHORT-NAME>SIGNALNAME8BIT</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitPerSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitPerSignallist/Address</DEFINITION-REF>\n\
                  <VALUE>ADDRESS8BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitPerSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>VALUE8BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg8bitPerSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>MASK8BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'
DataToEdit16BitPer ='<ECUC-CONTAINER-VALUE UUID="224569eb-a267-46de-b306-e4db22625bf1">\n\
              <SHORT-NAME>SIGNALNAME16BIT</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitPerSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitPerSignallist/Address</DEFINITION-REF>\n\
                  <VALUE>ADDRESS16BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitPerSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>VALUE16BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg16bitPerSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>MASK16BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'
DataToEdit32BitPer ='<ECUC-CONTAINER-VALUE UUID="224569eb-a267-46de-b306-e4db22625bf1">\n\
              <SHORT-NAME>SIGNALNAME32BIT</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitPerSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitPerSignallist/Address</DEFINITION-REF>\n\
                  <VALUE>ADDRESS32BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitPerSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>VALUE32BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfyCritReg32bitPerSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>MASK32BIT</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'  

DataToEditSysPer ='<ECUC-CONTAINER-VALUE UUID="58ec9575-ec56-4227-935f-77909b7565fb">\n\
              <SHORT-NAME>SIGNALNAME</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegPerSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegPerSignallist/RegisterID</DEFINITION-REF>\n\
                  <VALUE>SYSREGREGISTERID</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegPerSignallist/SelectionID</DEFINITION-REF>\n\
                  <VALUE>SYSREGSELECTIONID</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegPerSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>SYSREGVALUE</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegPerSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>SYSREGMASK</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'  

DataToEditSysInit ='<ECUC-CONTAINER-VALUE UUID="58ec9575-ec56-4227-935f-77909b7565fb">\n\
              <SHORT-NAME>SIGNALNAME</SHORT-NAME>\n\
              <DEFINITION-REF DEST="ECUC-PARAM-CONF-CONTAINER-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegInitSignallist</DEFINITION-REF>\n\
              <PARAMETER-VALUES>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegInitSignallist/RegisterID</DEFINITION-REF>\n\
                  <VALUE>SYSREGREGISTERID</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegInitSignallist/SelectionID</DEFINITION-REF>\n\
                  <VALUE>SYSREGSELECTIONID</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegInitSignallist/Value</DEFINITION-REF>\n\
                  <VALUE>SYSREGVALUE</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
                <ECUC-NUMERICAL-PARAM-VALUE>\n\
                  <DEFINITION-REF DEST="ECUC-INTEGER-PARAM-DEF">/Nxtr/VrfyCritReg/VrfySysCritRegInitSignallist/Mask</DEFINITION-REF>\n\
                  <VALUE>SYSREGMASK</VALUE>\n\
                </ECUC-NUMERICAL-PARAM-VALUE>\n\
              </PARAMETER-VALUES>\n\
            </ECUC-CONTAINER-VALUE>\n'
            
DataToAppendEnd =" </CONTAINERS>\n\
        </ECUC-MODULE-CONFIGURATION-VALUES>\n\
      </ELEMENTS>\n\
    </AR-PACKAGE>\n\
  </AR-PACKAGES>\n\
</AUTOSAR>\n"


        
outputpath=getfilename()
wb = load_workbook(outputpath,data_only=True)
sheet = wb.get_sheet_by_name('RH850-P1M')
a= sheet.max_row
b=sheet.max_column
n=2
data={}
for i in range(2,a+2):
    data[str(sheet.cell(row=i,column=2).value)]=[str(sheet.cell(row=i,column=15).value).replace(" ",""),str(sheet.cell(row=i,column=16).value).replace(" ",""),
                                                 str(sheet.cell(row=i,column=50).value).replace(" ",""),str(sheet.cell(row=i,column=49).value).replace(" ","").replace("H",""),
                                                 str(sheet.cell(row=i,column=13).value),str(sheet.cell(row=i,column=14).value).replace(" ","")]

for val in data.keys():
    if (((data[val][4]).upper().strip() == 'FALSE') or(((data[val][4]).upper().strip()==""))):
        del data[val]
print len(data.keys())
periodic8bit={}
periodic16bit={}
periodic32bit={}
InitOnly8bit={}
InitOnly16bit={}
InitOnly32bit={}
NrOfSysCritInitReg={}
NrOfSysCritPerReg ={}


for val in data.keys():
    a= data[val]
    if ("" not in a):        
        Address='0x'+ (a[3].strip().replace("0x",""))
        Value=a[1].strip().replace("0x","").replace("U","").upper()
        Value="0x"+Value
        if(a[0]=="N/A"):
            
            if a[2].strip() == '8':
                Mask="0xFF"
            elif a[2] == '16':
                Mask="0xFFFF"
            elif a[2] =='32':
                Mask="0xFFFFFFFF"
            else:
                print "Not Enough info for Mask"+ Address
                Mask =""
        else:
            Mask = "0x"+a[0].strip().replace("0x","").strip().replace("U","").replace("u","")
       
        if(a[5]=="Periodic"):
            if a[2].strip() == '8':
                periodic8bit[Address]=[str(int(Address,0)), str(int(Value, 0)), str(int(Mask, 0)), val]
            elif a[2].strip() == '16':
                periodic16bit[Address]=[str(int(Address,0)), str(int(Value, 0)), str(int(Mask, 0)), val]
            else:
                periodic32bit[Address]=[str(int(Address,0)), str(int(Value, 0)), str(int(Mask, 0)), val]

        elif ((a[5]=="InitOnly") or(a[4]=="Init Only")):
            if a[2].strip() == '8':
                InitOnly8bit[Address]=[str(int(Address,0)), str(int(Value, 0)), str(int(Mask, 0)), val]
            elif a[2].strip() == '16':
                InitOnly16bit[Address]=[str(int(Address,0)), str(int(Value, 0)), str(int(Mask, 0)), val]
            else:
                InitOnly32bit[Address]=[str(int(Address,0)), str(int(Value, 0)), str(int(Mask, 0)), val]
        else:
            print "Error in the Register " + val
    else:
        print "Error in the Register " + val 

sheet = wb.get_sheet_by_name('RH850-P1M CPU Core')
a= sheet.max_row
b=sheet.max_column
n=2
data={}
NrOfSysCritPerReg={}
NrOfSysCritInitReg={}
for i in range(2,a+2):
    data[str(sheet.cell(row=i,column=2).value)]=[str(sheet.cell(row=i,column=1).value).replace(" ",""),str(sheet.cell(row=i,column=20).value).replace(" ",""),
                                                 str(sheet.cell(row=i,column=25).value).replace(" ",""),str(sheet.cell(row=i,column=26).value).replace(" ",""),
                                                 str(sheet.cell(row=i,column=27).value)]

for val in data.keys():
    if ((data[val][1]).upper().strip().upper() != 'TRUE'):
        del data[val]


for val in data.keys():
    a= data[val]
    if ("" not in a):        
        RegisterID=((a[0].strip().replace("0x","")).split(","))[0].upper().replace("SR","")
        SelectionID=(a[0].strip().replace("0x","")).split(",")[1]
        Value="0x"+(a[3].strip().replace("0x",""))
        if(a[4]=="N/A"):
            Mask="0xFFFFFFFFU"
        else:
            Mask = "0x"+a[4].strip().replace("0x","").strip().replace("U","").replace("u","")
       
        if(a[2]=="Periodic"):
            NrOfSysCritPerReg[val]=[RegisterID,SelectionID, str(int(Value, 0)),  str(int(Mask, 0)), val]

        elif ((a[2]=="InitOnly") or(a[2]=="Init Only")):
            NrOfSysCritInitReg[val]=[RegisterID,SelectionID, str(int(Value, 0)),  str(int(Mask, 0)), val]
        else:
            print "Error in the Register " + val
    else:
        print "Error in the Register " + val
        
InitOnlyRegDef=''          
for value in InitOnly8bit.keys():
    InitOnlyRegDef= InitOnlyRegDef+ DataToEdit8BitInit.replace('ADDRESS8BIT' ,InitOnly8bit[value][0]).replace('VALUE8BIT' ,InitOnly8bit[value][1]).replace('MASK8BIT' ,InitOnly8bit[value][2]).replace('SIGNALNAME8BIT' ,InitOnly8bit[value][3])

for value in InitOnly16bit.keys():
    InitOnlyRegDef= InitOnlyRegDef+ DataToEdit16BitInit.replace('ADDRESS16BIT' ,InitOnly16bit[value][0]).replace('VALUE16BIT' ,InitOnly16bit[value][1]).replace('MASK16BIT' ,InitOnly16bit[value][2]).replace('SIGNALNAME16BIT' ,InitOnly16bit[value][3])

for value in InitOnly32bit.keys():
    InitOnlyRegDef= InitOnlyRegDef+ DataToEdit32BitInit.replace('ADDRESS32BIT' ,InitOnly32bit[value][0]).replace('VALUE32BIT' ,InitOnly32bit[value][1]).replace('MASK32BIT' ,InitOnly32bit[value][2]).replace('SIGNALNAME32BIT' ,InitOnly32bit[value][3])

periodicRegDef=''          
for value in periodic8bit.keys():
    periodicRegDef= periodicRegDef+ DataToEdit8BitPer.replace('ADDRESS8BIT' ,periodic8bit[value][0]).replace('VALUE8BIT' ,periodic8bit[value][1]).replace('MASK8BIT' ,periodic8bit[value][2]).replace('SIGNALNAME8BIT' ,periodic8bit[value][3])

for value in periodic16bit.keys():
    periodicRegDef= periodicRegDef+ DataToEdit16BitPer.replace('ADDRESS16BIT' ,periodic16bit[value][0]).replace('VALUE16BIT' ,periodic16bit[value][1]).replace('MASK16BIT' ,periodic16bit[value][2]).replace('SIGNALNAME16BIT' ,periodic16bit[value][3])

for value in periodic32bit.keys():
    periodicRegDef= periodicRegDef+ DataToEdit32BitPer.replace('ADDRESS32BIT' ,periodic32bit[value][0]).replace('VALUE32BIT' ,periodic32bit[value][1]).replace('MASK32BIT' ,periodic32bit[value][2]).replace('SIGNALNAME32BIT' ,periodic32bit[value][3])   
    
sysreg=' '
for value in NrOfSysCritPerReg.keys():
    sysreg = sysreg + DataToEditSysPer.replace("SYSREGREGISTERID",NrOfSysCritPerReg[value][0]).replace("SYSREGSELECTIONID",NrOfSysCritPerReg[value][1]).replace("SYSREGVALUE",NrOfSysCritPerReg[value][2]).replace("SYSREGMASK",NrOfSysCritPerReg[value][3]).replace("SIGNALNAME",NrOfSysCritPerReg[value][4])

for value in NrOfSysCritInitReg.keys():
    sysreg = sysreg + DataToEditSysInit.replace("SYSREGREGISTERID",NrOfSysCritInitReg[value][0]).replace("SYSREGSELECTIONID",NrOfSysCritInitReg[value][1]).replace("SYSREGVALUE",NrOfSysCritInitReg[value][2]).replace("SYSREGMASK",NrOfSysCritInitReg[value][3]).replace("SIGNALNAME",NrOfSysCritInitReg[value][4]) 

    
                                     
Sourcefile =    DataToAppendStart + InitOnlyRegDef+periodicRegDef+sysreg+DataToAppendEnd
f2=open('CDD_VrfyCritReg_Cfg.arxml', 'w')
f2.write(Sourcefile)
f2.close()

   
###print perRegDef   
##f1 = open('CDD_VrfyCritReg_Cfg_Template.h', 'r')
##Headerfile = f1.read()
##f1.close()
##Headerfile=Headerfile.replace("<Date>", strftime("%m-%d-%Y %H:%M:%S") );
##Headerfile=Headerfile.replace("<NrOfIninCritReg>",(str(len(InitOnly.keys()))+"U"))
##Headerfile=Headerfile.replace("<NrOfPerCritReg>", (str(len(periodic.keys()))+"U"))
##Headerfile=Headerfile.replace("<NrOfSysCritReg>", (str(len(NrOfSysCritReg.keys()))+"U"))
###print x1
##f1 = open('CDD_VrfyCritReg_Cfg.h', 'w')
##f1.write(Headerfile)
##f1.close()
##f2 = open('CDD_VrfyCritReg_Cfg_Template.c', 'r')
##Sourcefile = f2.read()
##f2.close()
##Sourcefile=Sourcefile.replace("<Date>", strftime("%m-%d-%Y %H:%M:%S") );
##Sourcefile=Sourcefile.replace("<InitCriticalregsiters>",InitOnlyRegDef)
##Sourcefile=Sourcefile.replace("<PerCriticalregsiters>",perRegDef)
##Sourcefile=Sourcefile.replace("<SysCriticalregsiters>",'{24, 4 , 0x00030003U , 0xFFFFFFFF},/*ICCTRL*/\n \t{26, 4 , 0x00010844U , 0xFFFFFFFF},/*ICCFG*/\n')
##
##f2=open('CDD_VrfyCritReg_Cfg.c', 'w')
##f2.write(Sourcefile)
##f2.close()
###print x2


